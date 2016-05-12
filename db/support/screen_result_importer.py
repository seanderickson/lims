from __future__ import unicode_literals

import argparse
import xlrd
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import logging
import json
from reports import ParseError, ValidationError
from db.support.data_converter import default_converter
from reports.serialize import parse_val
from collections import OrderedDict
from django.core.serializers.base import SerializationError
from xlsxwriter.utility import xl_col_to_name
logger = logging.getLogger(__name__)

from reports.serialize.xlsutils import sheet_cols, sheet_rows, \
    sheet_rows_by_colnames, workbook_sheets, generic_xls_write_workbook


PARTITION_POSITIVE_MAPPING = {
    'NP': 0,
    'W': 1,
    'M': 2,
    'S': 3 
}
CONFIRMED_POSITIVE_MAPPING = {
    'NT': 0,
    'I': 1,
    'FP': 2,
    'CP': 3 
}
ASSAY_WELL_CONTROL_TYPES = {
    'p': 'assay_positive_control',
    'n': 'assay_control',
    's': 'assay_control_shared',
    'o': 'other'
}
META_MAP = OrderedDict((
    ('Screen Number', 'screen_facility_id'),
))
DATA_COLUMN_FIELD_MAP = OrderedDict((
    ('"Data" Worksheet Column',  'data_worksheet_column'),
    ('name',  'name'),
    ('data type',  'data_type'),
    ('decimal places',  'decimal_places'),
    ('description',  'description'),
    ('replicate number',  'replicate_ordinal'),
    ('time point',  'time_point'),
    ('assay readout type',  'assay_readout_type'),
    ('if derived, how?',  'how_derived'),
    ('if derived, from which columns?',  'derived_from_columns'),
    ('primary or follow up?',  'is_follow_up_data'),
    ('which assay phenotype does it belong to?',  'assay_phenotype'),
    ('comments',  'comments'),
    ('channel', 'channel'),
    ('time point ordinal', 'time_point_ordinal'),
    ('zdepth ordinal', 'zdepth_ordinal'),
    ('clo id', 'clo_id'),
))
DATA_TYPES = [
    'text', 'numeric', 'partition_positive_indicator',
    'boolean_positive_indicator', 
    'confirmed_positive_indicator']
RESULT_VALUE_FIELD_MAP = OrderedDict((
    ('plate', 'plate_number'),
    ('well', 'well_name'),
    ('type', 'assay_well_control_type'),
    ('exclude', 'exclude'),
))
WELLNAME_MATCHER = re.compile(r'^[a-pA-P]{1,2}\d{1,2}$')


def data_column_field_mapper(fields):
    mapped_keys = []
    for key in fields:
        key = key.lower()
        mapped_key = None
        if key in DATA_COLUMN_FIELD_MAP:
            mapped_key = DATA_COLUMN_FIELD_MAP[key]
        else:
            for k in DATA_COLUMN_FIELD_MAP.values():
                if default_converter(key) == k:
                    mapped_key = k
                    break
        if not mapped_key:
            raise ParseError(
                key=key, 
                msg=('key %r is not in the recognized datacolumn fields: %r'
                    % (key,DATA_COLUMN_FIELD_MAP.keys())))
        mapped_keys.append(mapped_key)
    return mapped_keys

def parse_columns(columns_sheet):
    logger.info('parsing columns sheet: %r', columns_sheet.name)
    columns = data_column_generator(sheet_cols(columns_sheet))
    parsed_cols = OrderedDict()
    errors = {}
    for i,column in enumerate(columns):
        parsed_col = {
            'is_derived': False,
            'is_follow_up_data': False,
            'ordinal': i
        }
        logger.debug('parsing column: %r', column['data_worksheet_column'])
        parsed_cols[column['data_worksheet_column']] = parsed_col
        for key,val in column.items():
            if key == 'is_follow_up_data':
                parsed_col[key] = ( val and val.lower() == 'follow up')
            elif key == 'data_type':
                val = default_converter(val)
                # handle validation errors in the api
                if val not in DATA_TYPES:
                    key = '%s:%s' % (column['data_worksheet_column'],'data_type')
                    errors[key] = 'val: %r must be one of %r' % (val,DATA_TYPES)
                parsed_col[key] = val
            elif key == 'assay_readout_type':
                parsed_col[key] = default_converter(val)
            else:
                if key == 'how_derived':
                    parsed_col['is_derived'] = ( 
                        val is not None and val.strip() is not '' )
                parsed_col[key] = val
        
        if parsed_col.get('decimal_places', None):
            try:
                key = '%s:%s' % (column['data_worksheet_column'],'data_type')
                column['decimal_places'] = parse_val(
                    column['decimal_places'],key,'integer')
            except ValidationError, e:
                errors.update(e.errors)
        logger.debug('parsed_col: %r', parsed_col)
    if errors:
        raise ValidationError(errors={'Data Columns': errors})
    return parsed_cols
        
def result_value_field_mapper(header_row):
    logger.info('map result value header row...')
    mapped_row = []
    header_row = [x for x in header_row]
    for i,value in enumerate(header_row):
        found = False
        if value.lower() in RESULT_VALUE_FIELD_MAP:
            mapped_row.append(RESULT_VALUE_FIELD_MAP[value.lower()])
        else:
            mapped_row.append(xlrd.book.colname(i))
            
    unmapped = [key for key,value in RESULT_VALUE_FIELD_MAP.items() 
        if value not in mapped_row]
    if unmapped:
        msg=('missing fields: %r in result values header row: %r'
            % (str(','.join(unmapped)), header_row))
        logger.info(msg)
        raise ParseError(key='header row',msg=msg)
    logger.info('mapped result value header row: %r', mapped_row) 
    return mapped_row
        
def parse_result_values(parsed_columns, sheets):
    logger.info('parse result values...')
    well_ids = set()
    parse_error = None
    for sheet in sheets:
        logger.info('parse result values sheet: %r...', sheet.name)
    
        rows = sheet_rows(sheet)
        try:
            header_row = result_value_field_mapper(rows.next())
        except ValidationError, e:
            logger.info('error: %r', e)
            if not parse_error:
                parse_error = ParseError(errors={})
            if not sheet.name in parse_error.errors:
                parse_error.errors[str(sheet.name)] = {}
            parse_error.errors[sheet.name] = e.errors
            continue
        
        for i,row in enumerate(rows):
            try:
                result = parse_result_row(
                    i,parsed_columns,dict(zip(header_row,row)))
                if result['well_id'] in well_ids:
                    raise ParseError(
                        key=result['well_id'],
                        msg='duplicate')
                well_ids.add(result['well_id'])
                yield result
            except ValidationError,e:
                logger.warn('error: %r', e)
                if not parse_error:
                    parse_error = ParseError(errors={})
                if not sheet.name in parse_error.errors:
                    parse_error.errors[sheet.name] = {}
                parse_error.errors[sheet.name].update(e.errors)
#                 yield []
    if parse_error:
        raise parse_error
    
def parse_result_row(i,parsed_columns,result_row):    
    
    logger.debug('parse result row: %r', result_row)
    
    meta_columns = RESULT_VALUE_FIELD_MAP.values()
    parsed_row = {}
    excluded_cols = []
    
    meta_key = 'plate_number'
    val = result_row[meta_key]
    logger.debug('plate value to parse: %r', val)
    plate_number = parse_val(val, meta_key, 'integer')
    meta_key = 'well_name'
    val = result_row[meta_key]
    if WELLNAME_MATCHER.match(val):
        wellname = val
    else:
        raise ParseError(
            key=i, 
            msg=('well_name val %r does not follow the pattern: %r'
            % (val, WELLNAME_MATCHER.pattern))) 
    parsed_row['well_id'] = \
        '%s:%s' % (str(plate_number).zfill(5), wellname)
    
    meta_key = 'assay_well_control_type'
    val = result_row.get(meta_key, None)
    parsed_row[meta_key] = None
    if val:
        if val.lower() in ASSAY_WELL_CONTROL_TYPES:
            parsed_row[meta_key] = \
                ASSAY_WELL_CONTROL_TYPES[val.lower()]
        else:
            msg = ('%s: val %r is not one of the choices: %r'
                % (meta_key, val, ASSAY_WELL_CONTROL_TYPES))
            logger.error(msg)
            raise ValidationError(key=parsed_row['well_id'], msg=msg)

    meta_key = 'exclude'
    val = result_row.get(meta_key, None)
    if val:
        if val.lower() == 'all':
            excluded_cols = parsed_columns.keys()
        else:
            excluded_cols = [
                x.strip().upper() for x in val.split(',')]
            unknown_excluded_cols = (
                set(excluded_cols) - set(parsed_columns.keys()))
            if unknown_excluded_cols:
                raise ValidationError(
                    key = parsed_row['well_id'],
                    msg = 'unknown excluded cols: %r' % unknown_excluded_cols )
    
    for colname, raw_val in result_row.items():
        if colname in meta_columns:
            continue
        if colname not in parsed_columns:
            raise ValidationError(
                key='colname',
                msg='result value column is not in recognized columns: %r' 
                    % parsed_columns )
        column = parsed_columns[colname]
        if ( column['data_type'] == 'partition_positive_indicator'
            and not raw_val):
            raw_val = 'NP' 
        if ( column['data_type'] == 'confirmed_positive_indicator'
            and not raw_val):
            raw_val = 'NT' 
        if raw_val is None:
            continue
        
        key = '%s-%s' % (parsed_row['well_id'],colname)
        rv_initializer = {}
        parsed_row[colname] = rv_initializer
        
        if column['data_type'] == 'numeric':
            # decimal_places = parse_val(
            #     column['decimal_places'],'decimal_places','integer')
            if  column['decimal_places'] > 0:
                # parse, to validate
                parse_val(raw_val, key, 'float')
                # record the raw val, to save all digits (precision)
                rv_initializer['numeric_value'] = raw_val
            else:
                rv_initializer['numeric_value'] = \
                    parse_val(raw_val, key, 'integer')
        elif column['data_type'] == 'partition_positive_indicator':
            val = raw_val.upper()
            if val not in PARTITION_POSITIVE_MAPPING:
                raise ValidationError(
                    key=key, 
                    msg='val: %r must be one of %r'
                        % (raw_val, PARTITION_POSITIVE_MAPPING.keys()))
            rv_initializer['value'] = val
        elif column['data_type'] == 'confirmed_positive_indicator':
            val = raw_val.upper()
            if val not in CONFIRMED_POSITIVE_MAPPING:
                raise ValidationError(
                    key=key, 
                    msg='val: %r must be one of %r'
                        % (raw_val, CONFIRMED_POSITIVE_MAPPING.keys()))
            rv_initializer['value'] = val
        elif column['data_type'] == 'boolean_positive_indicator':
            val = parse_val(raw_val,key,'boolean')
            rv_initializer['value'] = val
        else:
            rv_initializer['value'] = raw_val
        
        if colname in excluded_cols:
            rv_initializer['is_exclude'] = True
        else:
            rv_initializer['is_exclude'] = False
        
        logger.debug('rv_initializer: %r', rv_initializer)
            
    return parsed_row

def data_column_generator(data_columns):
    fields = [x for x in data_column_field_mapper(data_columns.next())]
    for col_def in data_columns:
        yield dict(zip(fields,col_def))

def read_workbook(wb):
    try:
        logger.info('read screen result file sheets...')
        sheets = workbook_sheets(wb)
        sheet = sheets.next()
        logger.info('read Screen Info sheet: %r', sheet.name)
        meta_cols = sheet_cols(sheet)
        meta_raw = dict(zip(meta_cols.next(),meta_cols.next()))
        meta = { META_MAP.get(key, key):val for key,val in meta_raw.items() }
        
        logger.info('meta: %r', meta)
        
        sheet = sheets.next()
        logger.info('read Data Columns sheet: %r...', sheet.name)
        columns = parse_columns(sheet)
        
        result_values = parse_result_values(columns,sheets)
        return OrderedDict((
            ('meta', meta),
            ('fields', columns),
            ('objects', result_values),
            ))
    except ValidationError,e:
        logger.exception('parse error: %r', e)
        raise 
   
# def create_workbook(file, meta, fields, results ):
#     ####
#     REPLACED BY SERIALIZE TO DATA STRUCTURE
#     ####
#     wb = xlsxwriter.Workbook(file, {'constant_memory': True})                    
#     sheet = wb.add_worksheet('Screen Info')
#     for i,row in enumerate(dict_to_rows(meta)):
#         logger.info('meta row: %r', row)
#         sheet.write_row(i,0,row)
#     
#     sheet = wb.add_worksheet('Data Columns')
#     sheet_labels = DATA_COLUMN_FIELD_MAP.keys()
#     data_columns = [key for key,field in fields.items() 
#         if ( field.get('is_datacolumn',False) 
#             or field.get('data_worksheet_column', None))  ]
#     data_columns = sorted(data_columns, key=lambda x: fields[x]['ordinal'])
#     logger.debug('data columns: %r', data_columns)
#     header_row = [sheet_labels[0]]
#     header_row.extend([get_column_letter(len(RESULT_VALUE_FIELD_MAP)+i+1) 
#         for i in range(len(data_columns))])
#     logger.debug('header_row: %r', header_row)
#     sheet.write_row(0,header_row)
#     
#     for i,(sheet_label,sheet_key) in enumerate(
#             DATA_COLUMN_FIELD_MAP.items()[1:]):
#         row = [sheet_label]
#         for j,key in enumerate(data_columns):
#             val = fields[key].get(sheet_key, None)
#             if sheet_key == 'data_type':
#                 val = fields[key].get('assay_data_type',fields[key].get('data_type',None))
#             if val:
#                 if sheet_key == 'is_follow_up_data':
#                     if val == True:
#                         val = 'Follow up'
#                     elif val == False:
#                         val = 'Primary'
#                 row.append(str(val))
#             else:
#                 row.append(None)
#                 logger.debug(
#                     'Note: sheet key not found in schema field: %r, %r', 
#                     sheet_key, fields[key])
#         logger.debug('data column row: %r', row)
#         sheet.write_row(i+1,row)
#         
#         
#     logger.info('Write the result values sheet')
#     sheet = workbook.add_worksheet('Data')
#     header_row = []
#     header_row.extend(RESULT_VALUE_FIELD_MAP.keys())
#     header_row.extend([fields[key].get('title', key) for key in data_columns])
#     sheet.write_row(0,header_row)
# 
#     row_count = 0
#     for result_value in result_values:
#         row_count += 1
#         row = []
#         
#         row.extend(result_value['well_id'].split(':'))
#         if ( result_value.has_key('assay_well_control_type')
#              and result_value['assay_well_control_type'] ):
#             row.append(
#                 control_type_mapping[default_converter(
#                     result_value['assay_well_control_type'])])
#         else:
#             row.append(None)
#         excluded_cols = []
#         if result_value.has_key('excluded') and result_value['excluded']:
#             for data_column_name in result_value['excluded']:
#                 excluded_cols.append(get_column_letter(
#                     len(RESULT_VALUE_FIELD_MAP)+1
#                         +data_columns.index(data_column_name)))
#                 excluded_cols = sorted(excluded_cols)
#         row.append(','.join(excluded_cols))
#         
#         for j,key in enumerate(data_columns):
#             if result_value.has_key(key):
#                 result_value_entry = result_value[key]
#                 # Test the entry:
#                 # 1: directly serializing dicts read in from a deserialized workbook
#                 # 2: directly serializing output from the api
#                 if isinstance(result_value_entry, dict):
#                     if 'numeric_value' in result_value_entry:
#                         row.append(result_value_entry['numeric_value'])
#                     elif 'value' in result_value_entry:
#                         row.append(result_value_entry['value'])
#                     else:
#                         logger.warn(
#                             'no value entry found in the result_value: %d', 
#                             result_value)
#                         row.append(None)
#                 else:
#                     row.append(result_value[key])
#             else:
#                 row.append(None)
#         sheet.write_row(row_count,row)
#         
#         if row_count % 10000 == 0:
#             logger.info('wrote %d rows', row_count)
#         
#     logger.info('create_workbook done: result rows: %d', row_count)
#     wb.close()
#     return wb
#     
# def serialize_to_workbook(data):
# ####
# replaced by create_output_data
# 
# 
# #####
#     if 'meta' not in data:
#         raise SerializationError(
#             ('"meta" field required for serialization: '
#              'output: %r') % data)
#     if 'objects' not in data:
#         raise SerializationError('"objects" field required for serialization')
#     meta = data['meta']
#     result_values = data['objects']
#     if 'screen_facility_id' not in meta:
#         raise SerializationError(
#             '"screen_facility_id" field required in "meta" for serialization')
#     if 'fields' not in meta:
#         raise SerializationError(
#             '"fields" field required in "meta" for serialization')
#     screen_facility_id = meta['screen_facility_id']
#     fields = meta['fields']
#     
#     return create_workbook(screen_facility_id, fields, result_values)
    
def write_workbook(file, screen_facility_id, fields, result_values):
    generic_xls_write_workbook(
        file, 
        create_output_data(screen_facility_id, fields, result_values ))
    
def create_output_data(screen_facility_id, fields, result_values ):
    '''
    Translate Screen Result data into a data structure ready for Serialization:
    {
       'Screen Info': [ [ row1 ], [ row2 ]...].
       'Data Columns': [ [ row1 ], [ row2 ]...].
       'Data': [ [ row1 ], [ row2 ]...].
    }
    '''
    logger.info('create screen result data structure for %r', screen_facility_id)
    control_type_mapping = {v:k for k,v in ASSAY_WELL_CONTROL_TYPES.items()}

    data = OrderedDict()
    
    data['Screen Info'] = [ ('Screen Number', screen_facility_id) ]
    
    data_column_structure = []
    data['Data Columns'] = data_column_structure
    
    datacolumn_labels = DATA_COLUMN_FIELD_MAP.keys()
    data_columns = [key for key,field in fields.items() 
        if ( field.get('is_datacolumn',False) 
            or field.get('data_worksheet_column', None))  ]
    data_columns = sorted(data_columns, key=lambda x: fields[x]['ordinal'])
    logger.debug('data columns: %r', data_columns)
    header_row = [datacolumn_labels[0]]
    header_row.extend([xl_col_to_name(len(RESULT_VALUE_FIELD_MAP)+i) 
        for i in range(len(data_columns))])
    logger.debug('header_row: %r', header_row)
    data_column_structure.append(header_row)

    for i,(sheet_label,sheet_key) in enumerate(
            DATA_COLUMN_FIELD_MAP.items()[1:]):
        row = [sheet_label]
        for j,key in enumerate(data_columns):
            val = fields[key].get(sheet_key, None)
            if sheet_key == 'data_type':
                val = fields[key].get('assay_data_type',fields[key].get('data_type',None))
            if val:
                if sheet_key == 'is_follow_up_data':
                    if val == True:
                        val = 'Follow up'
                    elif val == False:
                        val = 'Primary'
                row.append(str(val))
            else:
                row.append(None)
                logger.debug(
                    'Note: sheet key not found in schema field: %r, %r', 
                    sheet_key, fields[key])
        logger.debug('data column row: %r', row)
        data_column_structure.append(row)
        
    def result_value_generator(result_values):
        
        logger.info('Write the result values sheet')
        header_row = []
        header_row.extend(RESULT_VALUE_FIELD_MAP.keys())
        header_row.extend([fields[key].get('title', key) for key in data_columns])
        yield(header_row)

        row_count = 0
        for result_value in result_values:
            row_count += 1
            row = []
            
            row.extend(result_value['well_id'].split(':'))
            if ( result_value.has_key('assay_well_control_type')
                 and result_value['assay_well_control_type'] ):
                row.append(
                    control_type_mapping[default_converter(
                        result_value['assay_well_control_type'])])
            else:
                row.append(None)
            excluded_cols = []
            if result_value.has_key('excluded') and result_value['excluded']:
                temp = result_value['excluded']
                if hasattr(temp, 'split'):
                    temp = temp.split(LIST_DELIMITER_SQL_ARRAY)
                    
                for data_column_name in temp:
                    excluded_cols.append(get_column_letter(
                        len(RESULT_VALUE_FIELD_MAP)+1
                            +data_columns.index(data_column_name)))
                    excluded_cols = sorted(excluded_cols)
            row.append(','.join(excluded_cols))
            
            for j,key in enumerate(data_columns):
                if result_value.has_key(key):
                    result_value_entry = result_value[key]
                    # Test the entry:
                    # 1: directly serializing dicts read in from a deserialized workbook
                    # 2: directly serializing output from the api
                    if isinstance(result_value_entry, dict):
                        if 'numeric_value' in result_value_entry:
                            row.append(result_value_entry['numeric_value'])
                        elif 'value' in result_value_entry:
                            row.append(result_value_entry['value'])
                        else:
                            logger.warn(
                                'no value entry found in the result_value: %d', 
                                result_value)
                            row.append(None)
                    else:
                        row.append(result_value[key])
                else:
                    row.append(None)
            if row_count % 10000 == 0:
                logger.info('wrote %d rows', row_count)
            yield row
    
    data['Data'] = result_value_generator(result_values)

    return data

    
def read_file(input_file):
    '''Read a serialized (xlsx) screen result:
    - sheets:
       "Screen Info" - two columns: A - attribute, B - value
       "Data Columns 
           - A - data column attribute name
           - [B-n] - data columns 
       "Data Sheets: old format ["PL_0001", "PL_0002, etc ]
       - new format: one sheet with a plate number column
    '''
    # FIXME: use openpyxl for memory optimization
    # from openpyxl import load_workbook
    # wb = load_workbook(filename='large_file.xlsx', read_only=True)
    # ws = wb['big_data'] # ws is now an IterableWorksheet
    # 
    # for row in ws.rows:
    #     for cell in row:
    #         print(cell.value)    
    
    
    logger.info('open screen result file for loading...')
    wb = xlrd.open_workbook(file_contents=input_file.read())
    return read_workbook(wb)

        
def json_printer(data):
    
    meta = data['meta']
    columns = meta['fields']
    result_values = data['objects']
    
    return ('{ "screen": %s, "columns": %s, "results": %s }'
        % ( json.dumps(meta['screen_facility_id']), 
            json.dumps(columns),
            json.dumps([result_value for result_value in result_values])
        ))

parser = argparse.ArgumentParser(
    description='reads a serialized (xlsx) screen result')

parser.add_argument(
    '-f', '--screen_result_file', required=True,
    help='''Screen Result file [meta, datacolumns, results].
            examples can be found in reports/static/test_data/screens''')
parser.add_argument(
    '-v', '--verbose', dest='verbose', action='count',
    help="Increase verbosity (specify multiple times for more)")    

if __name__ == "__main__":
    args = parser.parse_args()
    log_level = logging.WARNING # default
    if args.verbose == 1:
        log_level = logging.INFO
    elif args.verbose >= 2:
        log_level = logging.DEBUG
        DEBUG=True
    logging.basicConfig(
        level=log_level, 
        format='%(msecs)d:%(module)s:%(lineno)d:%(levelname)s: %(message)s')        

    with open(args.screen_result_file) as input_file:
        print json_printer(read_file(input_file))
        
        